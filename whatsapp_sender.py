#!/usr/bin/env python3
"""
WhatsApp Message Sender for Clinic Leads
Reads leads from leads.xlsx and sends personalized WhatsApp messages
via pywhatkit (WhatsApp Web automation).
"""

import time
import sys
import re
import argparse
import logging
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# ─── Configuration ────────────────────────────────────────────────────────────
EXCEL_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/leads.xlsx")
LOG_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/whatsapp_sender.log")
AIRBNB_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/airbnb_info.txt")

# Column indices (1-based, openpyxl)
COL_NAME = 1
COL_LOCATION = 2
COL_PHONE = 3
COL_EMAIL = 4
COL_CATEGORY = 5
COL_WEBSITE = 6
COL_DATE_ADDED = 7
COL_SCORE = 8
COL_MESSAGED = 9

# Timing defaults
DEFAULT_DELAY = 60          # seconds between messages
PYWHATKIT_WAIT_TIME = 15    # seconds for WhatsApp Web to load chat
PYWHATKIT_CLOSE_TIME = 5    # seconds before tab closes after sending

# Albania
COUNTRY_CODE_AL = "+355"

# ─── Airbnb Info Loader ──────────────────────────────────────────────────────
def load_airbnb_info():
    """
    Load Airbnb property details from airbnb_info.txt.
    Returns the file contents as a string, or None if file doesn't exist.

    Expected file format (plain text):
      Name: My Apartment in Tirana
      Location: Rruga e Dibres, Tirana
      Guests: 4
      Link: https://airbnb.com/rooms/12345
      (any other details you want included)
    """
    if not AIRBNB_FILE.exists():
        return None
    return AIRBNB_FILE.read_text(encoding="utf-8").strip()

# ─── Message Templates (Albanian) ────────────────────────────────────────────
MESSAGES = {
    "DENTAL": (
        "Pershendetje {name}!\n\n"
        "Ofrojme akomodim te pershtatshem per pacientet tuaj nderkombetare "
        "qe vijne ne Tirane per trajtim dentar.\n\n"
        "{airbnb_info}\n\n"
        "Pacientet tuaj mund te qendrojne ne nje apartament te rehatshem dhe "
        "te pajisur ploterisht, afer klinikes suaj. Kjo do t'u ofronte atyre "
        "nje pervojë me te mire dhe do t'ju ndihmonte te terhiqni me shume "
        "paciente nderkombetare.\n\n"
        "A do te ishit te interesuar per nje bashkepunim? Mund t'ua rekomandoni "
        "pacienteve tuaj si akomodimin e tyre te preferuar.\n\n"
        "Me respekt"
    ),
    "HAIR_TRANSPLANT": (
        "Pershendetje {name}!\n\n"
        "Ofrojme akomodim te pershtatshem per klientet tuaj nderkombetare "
        "qe vijne ne Tirane per transplant flokesh.\n\n"
        "{airbnb_info}\n\n"
        "Klientet tuaj mund te qendrojne ne nje apartament te rehatshem dhe "
        "te pajisur ploterisht, afer klinikes suaj — ideal per periudhen e "
        "rikuperimit pas transplantit. Kjo do t'u ofronte atyre nje pervojë "
        "me te mire.\n\n"
        "A do te ishit te interesuar per nje bashkepunim? Mund t'ua rekomandoni "
        "klienteve tuaj si akomodimin e tyre te preferuar.\n\n"
        "Me respekt"
    ),
    "UNKNOWN": (
        "Pershendetje {name}!\n\n"
        "Ofrojme akomodim te pershtatshem per pacientet tuaj nderkombetare "
        "qe vijne ne Tirane per trajtim mjekesor.\n\n"
        "{airbnb_info}\n\n"
        "Pacientet tuaj mund te qendrojne ne nje apartament te rehatshem dhe "
        "te pajisur ploterisht, afer klinikes suaj.\n\n"
        "A do te ishit te interesuar per nje bashkepunim?\n\n"
        "Me respekt"
    ),
}

# ─── Logging ─────────────────────────────────────────────────────────────────
def setup_logging():
    logger = logging.getLogger("whatsapp_sender")
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    return logger

# ─── Phone Number Formatting ────────────────────────────────────────────────
def format_phone_number(raw_phone, albanian_only=True):
    """
    Convert raw phone to international format for WhatsApp.
    Returns None if invalid or unsendable (landlines, non-Albanian when filtered).

    Handles:
      +355682032462     -> +355682032462  (already international Albanian)
      +39 02 8718 7132  -> None           (Italian, skipped if albanian_only)
      069 701 5050      -> +355697015050  (local Albanian mobile)
      0699676398        -> +355699676398  (local Albanian mobile, no spaces)
      04 223 1564       -> None           (Tirana landline)
    """
    if not raw_phone:
        return None

    phone = str(raw_phone).strip()
    if not phone:
        return None

    # Keep only digits and leading +
    digits_with_plus = re.sub(r'[^\d+]', '', phone)

    # Already international format
    if digits_with_plus.startswith('+'):
        if albanian_only and not digits_with_plus.startswith(COUNTRY_CODE_AL):
            return None
        if len(digits_with_plus) < 10:
            return None
        return digits_with_plus

    # Pure digits from here
    digits = re.sub(r'\D', '', phone)

    # Albanian landline (04X...) — no WhatsApp
    if digits.startswith('04'):
        return None

    # Albanian mobile with leading 0 (06X..., 10 digits)
    if digits.startswith('0') and len(digits) == 10:
        return COUNTRY_CODE_AL + digits[1:]

    # Albanian mobile without leading 0 (6X..., 9 digits)
    if digits.startswith('6') and len(digits) == 9:
        return COUNTRY_CODE_AL + digits

    return None

# ─── Name Cleaning ───────────────────────────────────────────────────────────
def clean_name(raw_name):
    """
    Clean clinic name for message greeting.
    Strips suffixes like '| Tirana', '- Dental Clinic', etc.
    """
    if not raw_name:
        return "klinika"
    name = str(raw_name).strip()
    # Split on common separators and keep first part
    for sep in ['|', ' - ', ' – ', ' — ', ': ']:
        if sep in name:
            name = name.split(sep)[0].strip()
    # Remove trailing location info after comma
    if ',' in name:
        parts = name.split(',')
        # Only strip if the part after comma looks like a location
        tail = parts[-1].strip().lower()
        if tail in ('tirana', 'albania', 'tiranë', 'tirane'):
            name = ','.join(parts[:-1]).strip()
    # Truncate if too long
    if len(name) > 60:
        name = name[:57] + "..."
    return name

# ─── Excel Operations ────────────────────────────────────────────────────────
def load_leads():
    """Load all leads from Excel. Returns list of dicts."""
    if not EXCEL_FILE.exists():
        return []

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    leads = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        name = row[COL_NAME - 1].value if len(row) >= COL_NAME else None
        phone = row[COL_PHONE - 1].value if len(row) >= COL_PHONE else None
        category = row[COL_CATEGORY - 1].value if len(row) >= COL_CATEGORY else None
        messaged = row[COL_MESSAGED - 1].value if len(row) >= COL_MESSAGED else None

        leads.append({
            "row": row_idx,
            "name": name,
            "phone": phone,
            "category": category or "UNKNOWN",
            "messaged": messaged,
        })

    wb.close()
    return leads


def get_sendable_leads(leads, category_filter=None):
    """
    Filter leads to those that:
    1. Have a phone number
    2. Have not been messaged yet
    3. Have a valid Albanian mobile number
    4. Match category filter (if set)
    """
    sendable = []
    for lead in leads:
        # Skip already messaged
        if lead["messaged"]:
            continue
        # Skip no phone
        if not lead["phone"]:
            continue
        # Category filter
        if category_filter and category_filter != "ALL":
            if lead["category"] != category_filter:
                continue
        # Format phone
        formatted = format_phone_number(lead["phone"])
        if not formatted:
            continue

        lead["formatted_phone"] = formatted
        sendable.append(lead)

    return sendable


def mark_as_messaged(row_number):
    """Update Excel to mark a row as messaged with current timestamp."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Ensure header exists
    header_cell = ws.cell(row=1, column=COL_MESSAGED)
    if not header_cell.value:
        header_cell.value = "Messaged"

    # Write timestamp
    ws.cell(row=row_number, column=COL_MESSAGED,
            value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    wb.save(EXCEL_FILE)
    wb.close()

# ─── Message Sending ────────────────────────────────────────────────────────
def send_message(phone, message, dry_run=False):
    """
    Send a WhatsApp message via pywhatkit.
    Returns True on success, False on failure.
    """
    if dry_run:
        return True

    try:
        import pywhatkit
        pywhatkit.sendwhatmsg_instantly(
            phone_no=phone,
            message=message,
            wait_time=PYWHATKIT_WAIT_TIME,
            tab_close=True,
            close_time=PYWHATKIT_CLOSE_TIME,
        )
        return True
    except Exception as e:
        logging.getLogger("whatsapp_sender").error(f"Failed to send to {phone}: {e}")
        return False

# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Send WhatsApp messages to clinic leads")
    parser.add_argument("--dry-run", "-d", action="store_true",
                        help="Simulate sending without actually calling pywhatkit")
    parser.add_argument("--limit", "-l", type=int, default=0,
                        help="Maximum number of messages to send (0 = no limit)")
    parser.add_argument("--delay", type=int, default=DEFAULT_DELAY,
                        help=f"Seconds between messages (default: {DEFAULT_DELAY})")
    parser.add_argument("--category", "-c",
                        choices=["DENTAL", "HAIR_TRANSPLANT", "UNKNOWN", "ALL"],
                        default="ALL",
                        help="Send only to a specific category (default: ALL)")
    args = parser.parse_args()

    logger = setup_logging()
    mode = "DRY RUN" if args.dry_run else "LIVE"

    print()
    print("═══════════════════════════════════════════════════")
    print("  WHATSAPP MESSAGE SENDER")
    print(f"  Mode: {mode}")
    print(f"  File: {EXCEL_FILE}")
    if args.limit:
        print(f"  Limit: {args.limit} messages")
    if args.category != "ALL":
        print(f"  Category: {args.category}")
    print(f"  Delay: {args.delay}s between messages")
    print("═══════════════════════════════════════════════════")
    print()

    # Load Airbnb info
    airbnb_info = load_airbnb_info()
    if airbnb_info:
        print(f"  Airbnb info: loaded from {AIRBNB_FILE.name}")
    else:
        print(f"  WARNING: {AIRBNB_FILE} not found!")
        print(f"  Create this file with your Airbnb property details.")
        print(f"  Messages will be sent without Airbnb details.")
        airbnb_info = "(Detajet e apartamentit do te shtohen se shpejti)"
    print()

    # Load and filter leads
    all_leads = load_leads()
    with_phone = [l for l in all_leads if l["phone"]]
    already_messaged = [l for l in all_leads if l["messaged"]]
    sendable = get_sendable_leads(all_leads, category_filter=args.category)

    print(f"  Total leads:        {len(all_leads)}")
    print(f"  With phone:         {len(with_phone)}")
    print(f"  Already messaged:   {len(already_messaged)}")
    print(f"  Sendable:           {len(sendable)}")

    if args.limit:
        sendable = sendable[:args.limit]
        print(f"  To send this run:   {len(sendable)} (limit)")
    else:
        print(f"  To send this run:   {len(sendable)}")
    print()

    if not sendable:
        print("  No leads to send messages to.")
        return

    # Sending loop
    sent = 0
    failed = 0

    for idx, lead in enumerate(sendable, start=1):
        name = clean_name(lead["name"])
        phone = lead["formatted_phone"]
        category = lead["category"]
        template = MESSAGES.get(category, MESSAGES["UNKNOWN"])
        message = template.format(name=name, airbnb_info=airbnb_info)

        print(f"  [{idx}/{len(sendable)}] {name} | {phone} | {category}")

        if args.dry_run:
            print(f"        [DRY RUN] Would send message ({len(message)} chars)")
            logger.info(f"DRY RUN | {phone} | {name} | {category}")
            sent += 1
        else:
            logger.info(f"Sending to {phone} | {name} | {category}")
            success = send_message(phone, message)

            if success:
                mark_as_messaged(lead["row"])
                sent += 1
                logger.info(f"SUCCESS | {phone} | {name}")
                print(f"        Sent successfully")
            else:
                failed += 1
                print(f"        FAILED — see log for details")

        sys.stdout.flush()

        # Delay between messages (skip after last one)
        if idx < len(sendable):
            if not args.dry_run:
                print(f"        Waiting {args.delay}s...")
                time.sleep(args.delay)
            else:
                time.sleep(0.1)

        # Progress report every 10 messages
        if idx % 10 == 0:
            print(f"  ── Progress: {sent} sent, {failed} failed, "
                  f"{len(sendable) - idx} remaining")
            print()

    # Final summary
    print()
    print("═══════════════════════════════════════════════════")
    print("  COMPLETE")
    print(f"  Sent: {sent} | Failed: {failed}")
    if args.dry_run:
        print("  Mode: DRY RUN (no messages were actually sent)")
    print("═══════════════════════════════════════════════════")


if __name__ == "__main__":
    main()
