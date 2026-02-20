#!/usr/bin/env python3
"""
Email Scraper — sequential with signal.alarm hard timeout.
Guaranteed to never hang on any site.
"""

import re
import sys
import signal
import subprocess
from openpyxl import load_workbook
from pathlib import Path
from urllib.parse import urljoin, urlparse

EXCEL_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/leads.xlsx")
COL_NAME = 1
COL_EMAIL = 4
COL_WEBSITE = 6

IGNORE_DOMAINS = {
    "example.com", "sentry.io", "wixpress.com", "googleapis.com",
    "w3.org", "schema.org", "wordpress.org", "wordpress.com",
    "gravatar.com", "jetpack.com", "google.com", "facebook.com",
    "twitter.com", "instagram.com", "youtube.com",
}


class TimeoutError(Exception):
    pass


def timeout_handler(signum, frame):
    raise TimeoutError()


def extract_emails(html):
    if not html or len(html) > 500000:  # skip huge pages
        return []
    pattern = r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
    raw = re.findall(pattern, html[:200000])  # only search first 200KB
    valid = set()
    for e in raw:
        e = e.lower().strip().rstrip('.')
        if e.endswith(('.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.css', '.js', '.woff', '.ttf', '.map')):
            continue
        d = e.split('@')[1] if '@' in e else ''
        if d in IGNORE_DOMAINS:
            continue
        if len(e) > 60:
            continue
        valid.add(e)
    return list(valid)


def pick_best(emails):
    if not emails:
        return None
    unique = list(dict.fromkeys(emails))
    for pfx in ['info@', 'contact@', 'office@', 'clinic@', 'reception@']:
        for e in unique:
            if e.startswith(pfx):
                return e
    return unique[0]


def curl_fetch(url):
    """Fetch URL using curl. Hard-killed after 6 seconds total."""
    try:
        proc = subprocess.Popen(
            ["curl", "-sL", "--max-time", "5", "--connect-timeout", "3",
             "-k", "-H", "User-Agent: Mozilla/5.0", url],
            stdout=subprocess.PIPE, stderr=subprocess.DEVNULL
        )
        try:
            stdout, _ = proc.communicate(timeout=6)
            if proc.returncode == 0 and stdout:
                return stdout.decode('utf-8', errors='replace')
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
    except Exception:
        pass
    return None


def scrape_site(url):
    """Try main page, then contact link, then /contact. Max ~15s per site."""
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url

    # Main page
    html = curl_fetch(url)
    if html:
        emails = extract_emails(html)
        if emails:
            return pick_best(emails)

        # Contact link from page
        links = re.findall(
            r'href=["\']([^"\']*(?:contact|kontakt)[^"\']*)["\']', html, re.IGNORECASE
        )
        for link in links[:1]:
            full = urljoin(url, link)
            if full != url:
                ch = curl_fetch(full)
                if ch:
                    emails = extract_emails(ch)
                    if emails:
                        return pick_best(emails)

    # /contact fallback
    p = urlparse(url)
    base = f"{p.scheme}://{p.netloc}"
    ch = curl_fetch(base + "/contact")
    if ch:
        emails = extract_emails(ch)
        if emails:
            return pick_best(emails)

    return None


def main():
    print("═══════════════════════════════════════════════════")
    print("  EMAIL SCRAPER — sequential + signal.alarm")
    print("═══════════════════════════════════════════════════")
    print(flush=True)

    # Set up alarm signal handler
    signal.signal(signal.SIGALRM, timeout_handler)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    targets = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        name_val = row[COL_NAME - 1].value if len(row) >= COL_NAME else None
        email_val = row[COL_EMAIL - 1].value if len(row) >= COL_EMAIL else None
        website_val = row[COL_WEBSITE - 1].value if len(row) >= COL_WEBSITE else None
        if website_val and str(website_val).strip() and (not email_val or not str(email_val).strip()):
            targets.append({"row": row_idx, "name": name_val, "website": str(website_val).strip()})

    total = len(targets)
    print(f"  Targets: {total}", flush=True)
    print(flush=True)

    found = 0
    notfound = 0

    for idx, target in enumerate(targets, 1):
        short = (str(target["name"]) or "?")[:40]

        # Hard 20s alarm per site
        signal.alarm(20)
        try:
            email = scrape_site(target["website"])
        except TimeoutError:
            email = None
        except Exception:
            email = None
        finally:
            signal.alarm(0)  # cancel alarm

        if email:
            ws.cell(row=target["row"], column=COL_EMAIL, value=email)
            found += 1
            print(f"  [{idx:3d}/{total}] {short:<42s} -> {email}", flush=True)
        else:
            notfound += 1
            print(f"  [{idx:3d}/{total}] {short:<42s} -> (none)", flush=True)

        # Save every 25 rows
        if idx % 25 == 0:
            wb.save(EXCEL_FILE)
            print(f"  --- saved ({found} found so far) ---", flush=True)

    wb.save(EXCEL_FILE)
    wb.close()

    print()
    print("═══════════════════════════════════════════════════")
    print(f"  COMPLETE — Found: {found} | Not found: {notfound}")
    print(f"  Total: {total}")
    print("═══════════════════════════════════════════════════")


if __name__ == "__main__":
    main()
