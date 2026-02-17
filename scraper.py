#!/usr/bin/env python3
"""
Serper Clinic Lead Scraper - Page Iterator
For each query, iterates pages 1 through 100, saving all unique results to leads.xlsx.
Stops a query when 3 consecutive pages return only duplicates.
"""

import requests
import math
import time
import sys
from datetime import date
from openpyxl import load_workbook, Workbook
from pathlib import Path

# ─── Configuration ────────────────────────────────────────────────────────────
SERPER_API_KEY = "5a88736a239477affbba609346cd8a884da61248"
SERPER_URL = "https://google.serper.dev/places"

BASE_LAT = 41.3212598
BASE_LNG = 19.8230679

COUNTRY = "Albania"
COUNTRY_CODE = "al"
LOCAL_LANG = "sq"
CITY = "Tirana"
CITY_LOCAL = "Tiranë"

MAX_RESULTS_PER_QUERY = 20
START_PAGE = 100
MAX_PAGE = 200
EXCEL_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/leads.xlsx")

# ─── Queries by topic ────────────────────────────────────────────────────────
DENTAL_QUERIES = [
    "klinikë dentare",
    "dentist",
    "implant dentar",
    "ortodontist",
    "kirurgi orale",
    "dental clinic",
]

HAIR_QUERIES = [
    "transplant flokësh",
    "hair transplant clinic",
    "klinikë transplant flokësh",
]

# ─── Helpers ──────────────────────────────────────────────────────────────────
def haversine(lat1, lng1, lat2, lng2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

def proximity_score(clinic_lat, clinic_lng):
    if clinic_lat is None or clinic_lng is None:
        return 0
    dist_km = haversine(BASE_LAT, BASE_LNG, clinic_lat, clinic_lng)
    if dist_km <= 1: return 10
    elif dist_km <= 2: return 9
    elif dist_km <= 3: return 8
    elif dist_km <= 5: return 7
    elif dist_km <= 8: return 6
    elif dist_km <= 12: return 5
    elif dist_km <= 18: return 4
    elif dist_km <= 25: return 3
    elif dist_km <= 40: return 2
    else: return 1

def classify_category(query, place_category):
    query_lower = query.lower()
    cat_lower = (place_category or "").lower()
    dental_keywords = ["dental", "dentist", "dentar", "ortodont", "oral", "dentare"]
    hair_keywords = ["hair", "transplant", "flokë"]
    for kw in dental_keywords:
        if kw in query_lower or kw in cat_lower:
            return "DENTAL"
    for kw in hair_keywords:
        if kw in query_lower or kw in cat_lower:
            return "HAIR_TRANSPLANT"
    return "UNKNOWN"

def parse_place(place, query):
    lat = place.get("latitude")
    lng = place.get("longitude")
    phone = place.get("phoneNumber", "") or place.get("phone", "") or place.get("phone_number", "") or ""
    return {
        "name": place.get("title", ""),
        "address": place.get("address", ""),
        "phone": phone,
        "email": None,
        "category": classify_category(query, place.get("category", "")),
        "website": place.get("website", ""),
        "latitude": lat,
        "longitude": lng,
        "score": proximity_score(lat, lng),
    }

# ─── Excel operations ────────────────────────────────────────────────────────
def load_existing_leads():
    """Load existing names and addresses from Excel for dedup."""
    existing_names = set()
    existing_addresses = set()
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                existing_names.add(str(row[0]).strip().lower())
            if row[1]:
                existing_addresses.add(str(row[1]).strip().lower())
        wb.close()
    return existing_names, existing_addresses

def save_leads_to_excel(new_leads, existing_names, existing_addresses):
    """Append deduplicated leads to Excel. Returns (added_count, total_count)."""
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(["Name", "Location", "Phone", "Email", "Category", "Website", "Date Added", "Score"])

    added = 0
    for lead in new_leads:
        name_lower = lead["name"].strip().lower() if lead["name"] else ""
        addr_lower = lead["address"].strip().lower() if lead["address"] else ""

        if name_lower and name_lower in existing_names:
            continue
        if addr_lower and addr_lower in existing_addresses:
            continue

        ws.append([
            lead["name"],
            lead["address"] or f"{CITY}, {COUNTRY}",
            lead.get("phone", ""),
            lead.get("email", ""),
            lead["category"],
            lead.get("website", ""),
            str(date.today()),
            lead.get("score", 0),
        ])
        if name_lower:
            existing_names.add(name_lower)
        if addr_lower:
            existing_addresses.add(addr_lower)
        added += 1

    wb.save(EXCEL_FILE)
    total = ws.max_row - 1
    wb.close()
    return added, total

def update_missing_phones(leads):
    """Update phone numbers for existing rows that have empty phones."""
    if not EXCEL_FILE.exists():
        return 0
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    name_to_row = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0]:
            name_lower = str(row[0]).strip().lower()
            phone_val = row[2] if len(row) > 2 else None
            if not phone_val or str(phone_val).strip() == "":
                name_to_row[name_lower] = row_idx

    updated = 0
    for lead in leads:
        phone = lead.get("phone", "")
        if not phone or str(phone).strip() == "":
            continue
        name_lower = lead["name"].strip().lower() if lead["name"] else ""
        if name_lower in name_to_row:
            ws.cell(row=name_to_row[name_lower], column=3, value=str(phone))
            updated += 1
            del name_to_row[name_lower]

    if updated > 0:
        wb.save(EXCEL_FILE)
    wb.close()
    return updated

# ─── Serper API call ──────────────────────────────────────────────────────────
def run_query(query, page=1):
    """Single Serper Places API call. Returns (places_list, error_string_or_None)."""
    payload = {
        "q": query,
        "location": "Tirana County, Albania",
        "gl": COUNTRY_CODE,
        "hl": LOCAL_LANG,
        "num": MAX_RESULTS_PER_QUERY,
        "ll": f"@{BASE_LAT},{BASE_LNG},14z",
        "page": page,
    }
    headers = {
        "X-API-KEY": SERPER_API_KEY,
        "Content-Type": "application/json",
    }
    try:
        resp = requests.post(SERPER_URL, json=payload, headers=headers, timeout=30)
        if resp.status_code in (403, 429):
            return [], f"API error {resp.status_code}: Credits exhausted"
        body = resp.text.lower()
        if any(kw in body for kw in ["credits", "limit exceeded", "exceeded"]):
            if resp.status_code != 200:
                return [], f"API credits exhausted (status {resp.status_code})"
        resp.raise_for_status()
        return resp.json().get("places", []), None
    except requests.exceptions.RequestException as e:
        return [], str(e)

# ─── Main: iterate pages 1-100 for each query ────────────────────────────────
def main():
    print("═══════════════════════════════════════════════════")
    print("  SERPER CLINIC LEAD SCRAPER — PAGE ITERATOR")
    print(f"  Location: {BASE_LAT}, {BASE_LNG} ({CITY}, {COUNTRY})")
    print(f"  Pages per query: {START_PAGE} → {MAX_PAGE}")
    print("═══════════════════════════════════════════════════")
    print()

    existing_names, existing_addresses = load_existing_leads()
    print(f"  Existing leads in file: {len(existing_names)} names")
    print()

    total_new_all = 0
    total_api_calls = 0
    credits_exhausted = False

    topics = [
        ("DENTAL", DENTAL_QUERIES),
        ("HAIR_TRANSPLANT", HAIR_QUERIES),
    ]

    for topic_name, queries in topics:
        if credits_exhausted:
            break

        print(f"═══════════════════════════════════════════════════")
        print(f"  TOPIC: {topic_name}")
        print(f"  Queries: {len(queries)}")
        print(f"═══════════════════════════════════════════════════")
        print()

        topic_new = 0

        for query in queries:
            if credits_exhausted:
                break

            print(f"  ┌─ Query: \"{query}\"")
            print(f"  │  Iterating pages {START_PAGE} → {MAX_PAGE}...")

            query_new = 0
            query_total_results = 0
            seen_titles_this_query = set()
            consecutive_dupe_pages = 0

            for page in range(START_PAGE, MAX_PAGE + 1):
                places, error = run_query(query, page=page)
                total_api_calls += 1

                if error:
                    if "403" in error or "429" in error or "credits" in error.lower() or "exhausted" in error.lower():
                        print(f"  │  Page {page}: *** API CREDITS EXHAUSTED ***")
                        credits_exhausted = True
                        break
                    print(f"  │  Page {page}: ERROR ({error}) — skipping")
                    continue

                if not places:
                    print(f"  │  Page {page}: empty — no more results for this query")
                    break

                # Check new vs dupes within this query
                new_on_page = 0
                dupes_on_page = 0
                page_leads = []

                for p in places:
                    title = (p.get("title") or "").strip().lower()
                    if title and title in seen_titles_this_query:
                        dupes_on_page += 1
                    else:
                        new_on_page += 1
                        if title:
                            seen_titles_this_query.add(title)
                        lead = parse_place(p, query)
                        page_leads.append(lead)

                query_total_results += new_on_page

                # Save this page's leads immediately
                added, total_in_file = save_leads_to_excel(page_leads, existing_names, existing_addresses)
                phones_updated = update_missing_phones(page_leads)

                query_new += added
                topic_new += added
                total_new_all += added

                status_parts = [f"{new_on_page} new, {dupes_on_page} dupes"]
                if added > 0:
                    status_parts.append(f"+{added} saved")
                if phones_updated > 0:
                    status_parts.append(f"+{phones_updated} phones updated")

                print(f"  │  Page {page:3d}: {' | '.join(status_parts)} | file: {total_in_file}")
                sys.stdout.flush()

                # Track consecutive pages with ALL duplicates
                if new_on_page == 0:
                    consecutive_dupe_pages += 1
                else:
                    consecutive_dupe_pages = 0

                # Stop this query if 3 consecutive pages had zero new results
                if consecutive_dupe_pages >= 3:
                    print(f"  │  → 3 consecutive dupe pages — moving to next query")
                    break

                time.sleep(0.3)

            print(f"  └─ Query done: {query_new} new leads from {query_total_results} unique results")
            print()

        print(f"  ── Topic {topic_name} done: +{topic_new} new leads")
        print()

    # Final summary
    print()
    print("═══════════════════════════════════════════════════")
    print("  SCRAPE COMPLETE")
    print(f"  Total API calls: {total_api_calls}")
    print(f"  Total new leads this run: {total_new_all}")
    if credits_exhausted:
        print(f"  Stop reason: API credits exhausted")
    else:
        print(f"  Stop reason: All queries exhausted")
    print(f"  File: {EXCEL_FILE}")
    print("═══════════════════════════════════════════════════")


if __name__ == "__main__":
    main()
