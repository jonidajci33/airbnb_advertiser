"""
Skill: Excel Dashboard Generator
Creates a detailed financial_dashboard.xlsx showing per-platform earnings,
expenses, and net by month/year/total. Includes a data tracker sheet.
"""

from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# --- Style constants ---
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SECTION_FONT = Font(bold=True, size=13, color="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
BOOKING_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
AIRBNB_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YEAR_TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GRAND_TOTAL_FILL = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
YEAR_TOTAL_FONT = Font(bold=True, size=11, color="1F3864")
GRAND_TOTAL_FONT = Font(bold=True, size=12, color="1F3864")
KPI_LABEL_FONT = Font(bold=True, size=11)
KPI_VALUE_FONT = Font(bold=True, size=14, color="1F3864")
OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CANCEL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
NOSHOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
EUR_FMT = '#,##0.00'
PCT_FMT = '0.0%'


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_w=10, max_w=22):
    for col_cells in ws.columns:
        max_len = min_w
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[col_letter].width = max_len


def _write_row(ws, row, values, fills=None, font=None, num_cols=None):
    """Write a row of values with optional styling."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = THIN_BORDER
        if fills:
            cell.fill = fills[col - 1] if isinstance(fills, list) else fills
        if font:
            cell.font = font
        if num_cols and col in num_cols:
            cell.number_format = EUR_FMT


# =============================================================================
# Sheet 1: Monthly Earnings (the main view the user wants)
# =============================================================================
def _create_monthly_earnings_sheet(wb, metrics):
    ws = wb.active
    ws.title = "Monthly Earnings"
    ws.sheet_properties.tabColor = "1F3864"

    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "Lulebore Apartment 1 - Monthly Earnings Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  All amounts in EUR"
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers - row 4
    row = 4
    headers = [
        "Month",
        "Booking.com\nGross", "Booking.com\nFees", "Booking.com\nNet",
        "Airbnb\nGross", "Airbnb\nFees", "Airbnb\nNet",
        "TOTAL\nGross", "TOTAL\nFees", "TOTAL\nNet",
        "Nights", "Bookings", "Occupancy", "ADR",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Sub-header colors for visual grouping
    # (Booking = blue, Airbnb = orange, Total = green)

    row = 5
    eur_cols = {2, 3, 4, 5, 6, 7, 8, 9, 10, 14}

    monthly = metrics["monthly"]
    yearly = metrics["yearly"]

    # Group months by year
    years_in_data = sorted(set(m["year"] for m in monthly.values()))

    for year in years_in_data:
        year_months = sorted(mk for mk, m in monthly.items() if m["year"] == year)

        for mk in year_months:
            m = monthly[mk]
            values = [
                m["month_label"],
                m["booking_gross"], m["booking_fees"], m["booking_net"],
                m["airbnb_gross"], m["airbnb_fees"], m["airbnb_net"],
                m["total_gross"], m["total_fees"], m["total_net"],
                m["total_nights"], m["total_bookings"],
                m["occupancy_pct"] / 100, m["adr"],
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                if col in eur_cols:
                    cell.number_format = EUR_FMT
                elif col == 13:
                    cell.number_format = PCT_FMT
                # Color code platform columns
                if col in (2, 3, 4):
                    cell.fill = BOOKING_FILL
                elif col in (5, 6, 7):
                    cell.fill = AIRBNB_FILL
                elif col in (8, 9, 10):
                    cell.fill = TOTAL_FILL
            row += 1

        # Yearly subtotal row
        if year in yearly:
            y = yearly[year]
            values = [
                f"TOTAL {year}",
                y["booking_gross"], y["booking_fees"], y["booking_net"],
                y["airbnb_gross"], y["airbnb_fees"], y["airbnb_net"],
                y["total_gross"], y["total_fees"], y["total_net"],
                y["total_nights"], y["total_bookings"],
                y["occupancy_pct"] / 100, y["adr"],
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.fill = YEAR_TOTAL_FILL
                cell.font = YEAR_TOTAL_FONT
                if col in eur_cols:
                    cell.number_format = EUR_FMT
                elif col == 13:
                    cell.number_format = PCT_FMT
            row += 1

    # Grand total row
    at = metrics["all_time"]
    values = [
        "GRAND TOTAL",
        at["booking_gross"], at["booking_fees"], at["booking_net"],
        at["airbnb_gross"], at["airbnb_fees"], at["airbnb_net"],
        at["total_gross"], at["total_fees"], at["total_net"],
        at["total_nights"], at["total_bookings"],
        at["occupancy_pct"] / 100, at["adr"],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = THIN_BORDER
        cell.fill = GRAND_TOTAL_FILL
        cell.font = GRAND_TOTAL_FONT
        if col in eur_cols:
            cell.number_format = EUR_FMT
        elif col == 13:
            cell.number_format = PCT_FMT

    chart_start = row + 3

    # Revenue chart
    # Find data range (skip yearly subtotal and grand total rows)
    data_rows = []
    for r in range(5, row):
        val = ws.cell(row=r, column=1).value
        if val and not str(val).startswith("TOTAL") and not str(val).startswith("GRAND"):
            data_rows.append(r)

    if len(data_rows) > 2:
        chart = LineChart()
        chart.title = "Monthly Net Earnings by Platform (EUR)"
        chart.y_axis.title = "EUR"
        chart.style = 10
        chart.width = 35
        chart.height = 15

        # We need to build the chart from non-contiguous rows, so use a simpler approach
        # Just reference the full range and it'll include subtotals (OK for visual)
        booking_net = Reference(ws, min_col=4, min_row=4, max_row=row - 1)
        airbnb_net = Reference(ws, min_col=7, min_row=4, max_row=row - 1)
        total_net = Reference(ws, min_col=10, min_row=4, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=5, max_row=row - 1)

        chart.add_data(booking_net, titles_from_data=True)
        chart.add_data(airbnb_net, titles_from_data=True)
        chart.add_data(total_net, titles_from_data=True)
        chart.set_categories(cats)

        # Style the series
        for s in chart.series:
            s.graphicalProperties.line.width = 22000

        ws.add_chart(chart, f"A{chart_start}")

    _auto_width(ws, min_w=12, max_w=18)
    ws.column_dimensions["A"].width = 16


# =============================================================================
# Sheet 2: Yearly Summary
# =============================================================================
def _create_yearly_summary_sheet(wb, metrics):
    ws = wb.create_sheet("Yearly Summary")
    ws.sheet_properties.tabColor = "BF8F00"

    ws.merge_cells("A1:N1")
    ws["A1"] = "Yearly Summary"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    headers = [
        "Year",
        "Booking.com\nGross", "Booking.com\nFees", "Booking.com\nNet",
        "Airbnb\nGross", "Airbnb\nFees", "Airbnb\nNet",
        "TOTAL\nGross", "TOTAL\nFees", "TOTAL\nNet",
        "Nights", "Bookings", "Occupancy", "YoY Growth",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(headers))

    eur_cols = {2, 3, 4, 5, 6, 7, 8, 9, 10}
    row = 4
    for year in sorted(metrics["yearly"].keys()):
        y = metrics["yearly"][year]
        values = [
            year,
            y["booking_gross"], y["booking_fees"], y["booking_net"],
            y["airbnb_gross"], y["airbnb_fees"], y["airbnb_net"],
            y["total_gross"], y["total_fees"], y["total_net"],
            y["total_nights"], y["total_bookings"],
            y["occupancy_pct"] / 100,
            y["yoy_growth_pct"] / 100 if y["yoy_growth_pct"] is not None else None,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if col in eur_cols:
                cell.number_format = EUR_FMT
            elif col in (13, 14):
                cell.number_format = PCT_FMT
            if col in (2, 3, 4):
                cell.fill = BOOKING_FILL
            elif col in (5, 6, 7):
                cell.fill = AIRBNB_FILL
            elif col in (8, 9, 10):
                cell.fill = TOTAL_FILL
        row += 1

    # Grand total
    at = metrics["all_time"]
    values = [
        "TOTAL",
        at["booking_gross"], at["booking_fees"], at["booking_net"],
        at["airbnb_gross"], at["airbnb_fees"], at["airbnb_net"],
        at["total_gross"], at["total_fees"], at["total_net"],
        at["total_nights"], at["total_bookings"],
        at["occupancy_pct"] / 100, None,
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = THIN_BORDER
        cell.fill = GRAND_TOTAL_FILL
        cell.font = GRAND_TOTAL_FONT
        if col in eur_cols:
            cell.number_format = EUR_FMT
        elif col == 13:
            cell.number_format = PCT_FMT

    # Bar chart
    if row > 5:
        chart = BarChart()
        chart.title = "Yearly Earnings by Platform (EUR)"
        chart.y_axis.title = "EUR"
        chart.style = 10
        chart.width = 22
        chart.height = 13

        data = Reference(ws, min_col=4, min_row=3, max_row=row - 1, max_col=4)
        data2 = Reference(ws, min_col=7, min_row=3, max_row=row - 1, max_col=7)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.add_data(data2, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{row + 2}")

    _auto_width(ws, min_w=12, max_w=18)


# =============================================================================
# Sheet 3: Dashboard KPIs
# =============================================================================
def _create_dashboard_sheet(wb, metrics):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "548235"

    at = metrics["all_time"]

    ws.merge_cells("A1:F1")
    ws["A1"] = "Key Performance Indicators"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Period: {at['date_range']} ({at['months_tracked']} months)"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:F2")

    row = 4
    kpis = [
        ("Total Gross Earnings", f"EUR {at['total_gross']:,.2f}"),
        ("Total Platform Fees", f"EUR {at['total_fees']:,.2f}"),
        ("Total Net Earnings", f"EUR {at['total_net']:,.2f}"),
        ("", ""),
        ("Booking.com Gross", f"EUR {at['booking_gross']:,.2f}"),
        ("Booking.com Fees", f"EUR {at['booking_fees']:,.2f}"),
        ("Booking.com Net", f"EUR {at['booking_net']:,.2f}"),
        ("", ""),
        ("Airbnb Gross", f"EUR {at['airbnb_gross']:,.2f}"),
        ("Airbnb Fees", f"EUR {at['airbnb_fees']:,.2f}"),
        ("Airbnb Net", f"EUR {at['airbnb_net']:,.2f}"),
        ("", ""),
        ("Avg Monthly Gross", f"EUR {at['avg_monthly_gross']:,.2f}"),
        ("Avg Monthly Net", f"EUR {at['avg_monthly_net']:,.2f}"),
        ("Overall Occupancy", f"{at['occupancy_pct']}%"),
        ("Overall ADR", f"EUR {at['adr']:,.2f}"),
        ("Total Room Nights", f"{at['total_nights']}"),
        ("Total Bookings", f"{at['total_bookings']}"),
        ("Unique Guests", f"{at['unique_guests']}"),
        ("Best Month", f"{at['best_month']['label']} (EUR {at['best_month']['gross']:,.2f})"),
        ("Worst Month", f"{at['worst_month']['label']} (EUR {at['worst_month']['gross']:,.2f})"),
    ]

    for label, value in kpis:
        if not label:
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = KPI_VALUE_FONT
        row += 1

    _auto_width(ws, min_w=20, max_w=45)


# =============================================================================
# Sheet 4: Booking.com Reservations
# =============================================================================
def _create_booking_details_sheet(wb, booking_reservations):
    ws = wb.create_sheet("Booking.com Details")
    ws.sheet_properties.tabColor = "2F5496"

    headers = [
        "Reservation #", "Status", "Arrival", "Departure", "Guest",
        "Nights", "Gross (EUR)", "Commission (EUR)", "Net (EUR)", "Month",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    sorted_res = sorted(booking_reservations, key=lambda r: r.get("arrival") or "9999")
    row = 2
    for r in sorted_res:
        values = [
            r["reservation_number"], r["status"],
            r["arrival"], r["departure"], r["guest_name"],
            r["room_nights"], r["gross_eur"], r["commission_eur"], r["net_eur"],
            r["month_key"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if col in (7, 8, 9):
                cell.number_format = EUR_FMT

        status = r["status"]
        if status == "CANCELLED":
            fill = CANCEL_FILL
        elif status == "NO_SHOW":
            fill = NOSHOW_FILL if not r["has_revenue"] else OK_FILL
        else:
            fill = OK_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    _auto_width(ws)


# =============================================================================
# Sheet 5: Airbnb Reservations
# =============================================================================
def _create_airbnb_details_sheet(wb, airbnb_reservations):
    ws = wb.create_sheet("Airbnb Details")
    ws.sheet_properties.tabColor = "C55A11"

    headers = [
        "Confirmation", "Arrival", "Departure", "Guest",
        "Nights", "Gross (USD)", "Fee (USD)", "Net (USD)",
        "Gross (EUR)", "Fee (EUR)", "Net (EUR)", "Rate", "Month",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    sorted_res = sorted(airbnb_reservations, key=lambda r: r.get("arrival") or "9999")
    row = 2
    for r in sorted_res:
        values = [
            r["reservation_number"],
            r["arrival"], r["departure"], r["guest_name"],
            r["room_nights"],
            r["gross_usd"], r["fee_usd"], r["net_usd"],
            r["gross_eur"], r["commission_eur"], r["net_eur"],
            r["exchange_rate"], r["month_key"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.fill = OK_FILL
            if col in (6, 7, 8, 9, 10, 11):
                cell.number_format = EUR_FMT
            elif col == 12:
                cell.number_format = '0.0000'
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    _auto_width(ws)


# =============================================================================
# Sheet 6: Data Tracker
# =============================================================================
def _create_tracker_sheet(wb, metrics, booking_months, airbnb_months):
    ws = wb.create_sheet("Data Tracker")
    ws.sheet_properties.tabColor = "7030A0"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Data Tracker - Months Processed"
    ws["A1"].font = TITLE_FONT

    ws["A2"] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")

    # All months in the data range
    all_months = sorted(metrics["months_processed"])

    row = 4
    headers = ["Month", "Booking.com", "Airbnb", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(headers))

    booking_set = set(booking_months)
    airbnb_set = set(airbnb_months)

    row = 5
    for mk in all_months:
        has_booking = mk in booking_set
        has_airbnb = mk in airbnb_set

        if has_booking and has_airbnb:
            status = "COMPLETE"
            fill = OK_FILL
        elif has_booking or has_airbnb:
            status = "PARTIAL"
            fill = NOSHOW_FILL
        else:
            status = "MISSING"
            fill = CANCEL_FILL

        values = [
            mk,
            "YES" if has_booking else "NO",
            "YES" if has_airbnb else "NO",
            status,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.fill = fill
        row += 1

    # Summary
    row += 1
    ws.cell(row=row, column=1, value="Total months tracked:").font = KPI_LABEL_FONT
    ws.cell(row=row, column=2, value=len(all_months))
    row += 1
    ws.cell(row=row, column=1, value="Booking.com months:").font = KPI_LABEL_FONT
    ws.cell(row=row, column=2, value=len(booking_set))
    row += 1
    ws.cell(row=row, column=1, value="Airbnb months:").font = KPI_LABEL_FONT
    ws.cell(row=row, column=2, value=len(airbnb_set))
    row += 1
    ws.cell(row=row, column=1, value="Next month to add:").font = KPI_LABEL_FONT
    # Figure out the next expected month
    if all_months:
        last = all_months[-1]
        y, m = int(last[:4]), int(last[5:7])
        m += 1
        if m > 12:
            m = 1
            y += 1
        ws.cell(row=row, column=2, value=f"{y}-{m:02d}")

    _auto_width(ws, min_w=14, max_w=20)


# =============================================================================
# Main entry point
# =============================================================================
def generate_excel(metrics, booking_reservations, airbnb_reservations,
                   booking_months, airbnb_months, output_path=None):
    """
    Generate the financial dashboard Excel file.

    Args:
        metrics: output from financial_calculator.compute_metrics()
        booking_reservations: list of Booking.com reservation dicts
        airbnb_reservations: list of Airbnb reservation dicts
        booking_months: list of month keys from Booking.com CSVs
        airbnb_months: list of month keys from Airbnb data
        output_path: optional path for the output file

    Returns:
        str: path to the generated Excel file
    """
    if output_path is None:
        output_path = PROJECT_ROOT / "financial_dashboard.xlsx"
    else:
        output_path = Path(output_path)

    wb = Workbook()

    _create_monthly_earnings_sheet(wb, metrics)
    _create_yearly_summary_sheet(wb, metrics)
    _create_dashboard_sheet(wb, metrics)
    _create_booking_details_sheet(wb, booking_reservations)
    _create_airbnb_details_sheet(wb, airbnb_reservations)
    _create_tracker_sheet(wb, metrics, booking_months, airbnb_months)

    wb.save(str(output_path))
    return str(output_path)
