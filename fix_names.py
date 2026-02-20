#!/usr/bin/env python3
"""
Fix clinic names in leads.xlsx.
Many names are webpage titles, not actual clinic names.
This script cleans them up.
"""

import re
from openpyxl import load_workbook
from pathlib import Path
from urllib.parse import urlparse

EXCEL_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/leads.xlsx")
COL_NAME = 1
COL_WEBSITE = 6

# Prefixes to strip (page title prefixes)
STRIP_PREFIXES = [
    "Contact Us En - ",
    "Contact Us - ",
    "Contact - ",
    "Kontakt - ",
    "Kontakt | ",
    "Na Kontaktoni - ",
    "Kush Eshte Klinika Dentare Me e Mire ne Tirane?",
]

# Suffixes to strip (after | or - or :)
STRIP_SUFFIXES = [
    "Home", "Tirana", "Albania", "Tiranë", "Tirane",
    "Dentisti in Albania", "Dentista in Albania", "Dentista Albania",
    "Dental Tourism", "Turismo Dentale", "Turismo Dentale in Albania",
    "Dental Clinic Tirana", "Dental Implants", "Dental Clinic",
    "Klinikë Dentare", "Klinike Dentare", "Klinika Dentare",
    "Dental Clinic in Tirana", "Dental Clinic in Tirana, Albania",
    "Clinica dentale Albania", "Clinica Dentale Albania",
    "Studio Dentistico a Tirana Albania ...",
    "Dental Hospital and Clinic Tirana Albania",
    "Compare Availability, Reviews ...",
    "Reviews", "Costs & Reviews",
    "Dentist near me", "Cosmetic dentist", "Dentist office",
    "Zahnartz", "Dental Aesthetics", "Implantology",
    "Implantology & Dental Aesthetics",
    "Advanced Implantology Center",
    "Dal Volo alla Cura per Denti e Sorrisi",
    "Affordable FUE & DHI in Tirana",
    "Affordable Prices & Results",
    "Turismo Dentale & Dentisti in Albania",
    "Dentisti in Albania",
    "Chirurgia estetica",
    "Impianti dentali",
    "Impianti Dentali Albania",
    "Clinica Dentale Odontoiatrica",
    "Aesthetic Dentist",
    "Klinike Dentare ne Tirane",
    "Expert Dental Care & Implants in Tirana, Albania",
    "Top-Rated Dentist in Albania",
    "Klinikë Dentare në Tiranë",
    "Klinika Dentare në Tiranë",
    "Digital Orthodontic Clinic",
    "Dental Day Hospital",
    "Klinikë Dentare Prof. Dr. Luan Mavriqi",
    "Dental Clinic and Lab in Tirana",
    "Turismo Dentale in Albania",
    "Clinica Dentale in Albania",
    "Dental Clinic in Durrës",
    "TURISMO DENTALE",
]

# Full name replacements (when the whole name is junk)
NAME_REPLACEMENTS = {
    "Top 10 Dental Clinics & Dentists in Tirana": "DentaVacation",
    "Dentists in Tirana – Best Clinics & Prices 2025": None,  # skip, not a clinic
    "Top 10+ Dentists in Tirana, Albania": None,
    "TOP 10+ Tooth Whitening doctors in Tirana, Costs & Reviews": None,
    "Best dentists in Tirana: TOP 16 doctors": None,
    "Best Clinics of Emergency Dentist in Albania, tirana": None,
    "Find 10 Orthodontics Clinics in Tirana": None,
    "DCA-Dental Center Albania Reviews 50": "Dental Center Albania",
    "Kirurg dentar dhe dentist - Tiranë": None,  # yellow pages listing
    "Klinika Dentare Tiranë": None,  # yellow pages listing
    "Dental • Klinike Dentare": None,  # directory listing
    "Dentisti in Albania - Dental Care Albania": "Dental Care Albania",
    "Kontakt - Klinike Dentare ne Tirane": None,  # generic contact page
    "Kontakt - VatanMed | Klinike Mbjellje Flokesh Ne Tirane": "VatanMed",
    "Kontakt | Finesa Dental | Klinikë Dentare në Tiranë": "Finesa Dental",
    "Kush Eshte Klinika Dentare Me e Mire ne Tirane?": "Emerald Dental",
    "What Is the Best Dental Clinic in Tirana, Albania?": "Virtus Dental Center",
    "Teeth Whitening in Oxa Clinic in Tirana.": "Oxa Clinic",
    "Teeth Whitening In Albania: Affordable Prices & Results": "City Dental Clinic CDC",
    "Cheap Dental Implants: the winning choice in Tirana, Albania": "Smile Provider",
    "Best Dental Implant Clinics in Albania": "Family Dental Care",
    "FUE Hair Transplant in Tirana": "Istanbul Care",
    "Hair Transplant in Albania Tirana": "Hair Solution Clinic",
    "Hair transplant in Albania | Advanced Hair Clinic in Tirana": "Advanced FUE Clinic",
    "Hair Transplant in Albania: Costs, Expert Clinics & Guided ...": "KEIT Clinic",
    "Hair Transplant Tirana with Micro FUE Technique": "DaVinci Clinic",
    "Hair Transplant in Albania - Affordable FUE & DHI in Tirana": "Estemoon Clinic",
    "Best hair transplant clinic in Albania for natural results": "Nouvelle Clinique",
    "KLINIKE PER MBJELLJE FLOKU – ADVANCED FUE CLINIC ...": "Advanced FUE Clinic",
    "Orthognathic surgery - Tiranë": "Salus Clinic",
    "Dental Center Albania Clinic - Compare Availability, Reviews ...": "Dental Center Albania",
    "Dental Clinic in Tirana - Orthosmile": "Orthosmile",
    "Dental Clinic in Tirana": "Orbis Dental Care",
    "Dental Clinic in Tirana - Queen Dental": "Queen Dental",
    "DR.BENARDA BEQARAJ - MJEKE STOMATOLOGE -DENTISTE": "Dr. Benarda Beqaraj",
    "Turismo Dentale Albania a Tirana": "Marco Licata Dental",
    "Turismo Dentale in Albania: Dent Dinamo AL": "Dent Dinamo",
    "Turismo Dentale in Albania, Dentisti Albania - Dental Care Albania": "Dental Care Albania",
    "Dental Tourism Albania - Dentale Albania": "Dentale Albania",
    "Dental Tourism Albania - Dental Clinic Albania": "Dental Care Albania",
    "San Marco Dental, Dentisti in Albania, Turismo Dentale": "San Marco Dental",
    "Massimo Sopranzi": "Massimo Sopranzi Dental",
    "Dental & Beauty | Chirurgia estetica | Impianti dentali": "Dental & Beauty",
    "TURISTA SORRIDENTE": "Turista Sorridente",
    "Albania Medical Tour": "Albania Medical Tour",
    "Spitali Nr 2 (Kirurgjia)": None,  # hospital, not a clinic partner
}


def clean_name(name, website=None):
    """Clean a clinic name."""
    if not name:
        return name

    name = str(name).strip()

    # Check full replacements first
    if name in NAME_REPLACEMENTS:
        return NAME_REPLACEMENTS[name]

    original = name

    # Strip known prefixes
    for prefix in STRIP_PREFIXES:
        if name.startswith(prefix):
            name = name[len(prefix):].strip()

    # Strip suffixes after separators (|, -, –, —, :)
    for sep in [' | ', ' - ', ' – ', ' — ', ': ']:
        if sep in name:
            parts = name.split(sep)
            # Try removing parts from the end that match known suffixes
            while len(parts) > 1:
                tail = parts[-1].strip()
                matched = False
                for suffix in STRIP_SUFFIXES:
                    if tail.lower() == suffix.lower() or tail.lower().startswith(suffix.lower()):
                        parts = parts[:-1]
                        matched = True
                        break
                if not matched:
                    break
            # Also try removing leading junk
            if len(parts) > 1:
                head = parts[0].strip()
                for suffix in STRIP_SUFFIXES:
                    if head.lower() == suffix.lower():
                        parts = parts[1:]
                        break
            name = parts[0].strip() if parts else name

    # Remove trailing dots and ellipsis
    name = name.rstrip('.').rstrip(' .').strip()
    if name.endswith('...'):
        name = name[:-3].strip()

    # Remove trailing year numbers like "2026", "2025"
    name = re.sub(r'\s+20\d{2}$', '', name)

    # Remove Durres/Albania/Tirana trailing after comma
    if ',' in name:
        parts = name.rsplit(',', 1)
        tail = parts[-1].strip().lower()
        if tail in ('tirana', 'albania', 'tiranë', 'tirane', 'durres', 'durrës'):
            name = parts[0].strip()

    # Remove parenthetical duplicates at the end
    # e.g., "Klinika Dentare Toti & Gurakuqi (Toti & Gurakuqi Dental Clinic)"
    paren_match = re.search(r'\s*\([^)]+\)\s*$', name)
    if paren_match and len(name) > 50:
        name = name[:paren_match.start()].strip()

    # B.E.L. Dent long suffix
    if 'B.E.L. Dent' in name and len(name) > 30:
        name = 'B.E.L. Dent'

    # Finesa Dental long
    if name.startswith('Finesa Dental') and len(name) > 30:
        name = 'Finesa Dental'

    return name if name else original


def main():
    print("═══════════════════════════════════════════════════")
    print("  NAME FIXER — Cleaning clinic names")
    print("═══════════════════════════════════════════════════")
    print()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    changed = 0
    removed = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        old_name = row[COL_NAME - 1].value
        website = row[COL_WEBSITE - 1].value if len(row) >= COL_WEBSITE else None

        if not old_name:
            continue

        new_name = clean_name(str(old_name), website)

        if new_name is None:
            # Mark as not useful but don't delete the row
            # Just leave name as is — the email sender will still work
            print(f"  SKIP  row {row_idx}: {str(old_name)[:60]}")
            removed += 1
            continue

        if new_name != str(old_name):
            ws.cell(row=row_idx, column=COL_NAME, value=new_name)
            changed += 1
            old_short = str(old_name)[:45]
            print(f"  FIX   row {row_idx}: {old_short:<47s} -> {new_name}")

    wb.save(EXCEL_FILE)
    wb.close()

    print()
    print("═══════════════════════════════════════════════════")
    print(f"  Names fixed: {changed}")
    print(f"  Rows flagged (not real clinics): {removed}")
    print("═══════════════════════════════════════════════════")


if __name__ == "__main__":
    main()
