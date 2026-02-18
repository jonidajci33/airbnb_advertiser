#!/usr/bin/env python3
"""
Email Sender for Clinic Leads
Reads leads from leads.xlsx and sends personalized emails
via Gmail SMTP. Follows the same message topology as whatsapp_sender.py.

Required environment variables:
  EMAIL_ADDRESS   — Gmail address (e.g. yourname@gmail.com)
  EMAIL_PASSWORD  — Gmail App Password (NOT your regular password)

To generate an App Password:
  1. Go to https://myaccount.google.com/apppasswords
  2. Select "Mail" and your device
  3. Copy the 16-character password
"""

import os
import time
import sys
import re
import argparse
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# ─── Configuration ────────────────────────────────────────────────────────────
EXCEL_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/leads.xlsx")
LOG_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/email_sender.log")
AIRBNB_FILE = Path("/mnt/windows/projekte Joni/airbnb_advertiser/airbnb_info.txt")

# Column indices (1-based, openpyxl)
COL_NAME = 1
COL_LOCATION = 2
COL_PHONE = 3
COL_EMAIL = 4
COL_CATEGORY = 5
COL_WEBSITE = 6
COL_DATE_ADDED = 7
COL_SCORE = 8
COL_MESSAGED = 9        # WhatsApp — not touched by this script
COL_EMAILED = 10        # Email — new column J

# Gmail SMTP
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

# Timing defaults
DEFAULT_DELAY = 10      # seconds between emails (Gmail allows ~20/sec but be safe)

# ─── Airbnb Info Loader ──────────────────────────────────────────────────────
def load_airbnb_info():
    """
    Load Airbnb property details from airbnb_info.txt.
    Returns the file contents as a string, or None if file doesn't exist.
    """
    if not AIRBNB_FILE.exists():
        return None
    return AIRBNB_FILE.read_text(encoding="utf-8").strip()

# ─── Message Templates (Albanian) — same as whatsapp_sender ──────────────────
MESSAGES = {
    "DENTAL": (
        "Pershendetje {name}!\n\n"
        "Ofrojme akomodim te pershtatshem per pacientet tuaj nderkombetare "
        "qe vijne ne Tirane per trajtim dentar.\n\n"
        "{airbnb_info}\n\n"
        "Pacientet tuaj mund te qendrojne ne nje apartament te rehatshem dhe "
        "te pajisur ploterisht, afer klinikes suaj. Kjo do t'u ofronte atyre "
        "nje pervojë me te mire dhe do t'ju ndihmonte te terhiqni me shume "
        "paciente nderkombetare.\n\n"
        "A do te ishit te interesuar per nje bashkepunim? Mund t'ua rekomandoni "
        "pacienteve tuaj si akomodimin e tyre te preferuar.\n\n"
        "Me respekt"
    ),
    "HAIR_TRANSPLANT": (
        "Pershendetje {name}!\n\n"
        "Ofrojme akomodim te pershtatshem per klientet tuaj nderkombetare "
        "qe vijne ne Tirane per transplant flokesh.\n\n"
        "{airbnb_info}\n\n"
        "Klientet tuaj mund te qendrojne ne nje apartament te rehatshem dhe "
        "te pajisur ploterisht, afer klinikes suaj — ideal per periudhen e "
        "rikuperimit pas transplantit. Kjo do t'u ofronte atyre nje pervojë "
        "me te mire.\n\n"
        "A do te ishit te interesuar per nje bashkepunim? Mund t'ua rekomandoni "
        "klienteve tuaj si akomodimin e tyre te preferuar.\n\n"
        "Me respekt"
    ),
    "UNKNOWN": (
        "Pershendetje {name}!\n\n"
        "Ofrojme akomodim te pershtatshem per pacientet tuaj nderkombetare "
        "qe vijne ne Tirane per trajtim mjekesor.\n\n"
        "{airbnb_info}\n\n"
        "Pacientet tuaj mund te qendrojne ne nje apartament te rehatshem dhe "
        "te pajisur ploterisht, afer klinikes suaj.\n\n"
        "A do te ishit te interesuar per nje bashkepunim?\n\n"
        "Me respekt"
    ),
}

# Email subject lines per category
SUBJECTS = {
    "DENTAL": "Bashkepunim — Akomodim per pacientet tuaj nderkombetare",
    "HAIR_TRANSPLANT": "Bashkepunim — Akomodim per klientet tuaj nderkombetare",
    "UNKNOWN": "Bashkepunim — Akomodim per pacientet tuaj",
}

# ─── Logging ─────────────────────────────────────────────────────────────────
def setup_logging():
    logger = logging.getLogger("email_sender")
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    return logger

# ─── Email Validation ────────────────────────────────────────────────────────
def validate_email(raw_email):
    """
    Basic email validation. Returns cleaned email or None if invalid.
    """
    if not raw_email:
        return None
    email = str(raw_email).strip().lower()
    if not email:
        return None
    # Basic regex: something@something.something
    if not re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email):
        return None
    return email

# ─── Name Cleaning (same as whatsapp_sender) ─────────────────────────────────
def clean_name(raw_name):
    """
    Clean clinic name for message greeting.
    Strips suffixes like '| Tirana', '- Dental Clinic', etc.
    """
    if not raw_name:
        return "klinika"
    name = str(raw_name).strip()
    # Split on common separators and keep first part
    for sep in ['|', ' - ', ' – ', ' — ', ': ']:
        if sep in name:
            name = name.split(sep)[0].strip()
    # Remove trailing location info after comma
    if ',' in name:
        parts = name.split(',')
        tail = parts[-1].strip().lower()
        if tail in ('tirana', 'albania', 'tiranë', 'tirane'):
            name = ','.join(parts[:-1]).strip()
    # Truncate if too long
    if len(name) > 60:
        name = name[:57] + "..."
    return name

# ─── Excel Operations ────────────────────────────────────────────────────────
def load_leads():
    """Load all leads from Excel. Returns list of dicts."""
    if not EXCEL_FILE.exists():
        return []

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    leads = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        name = row[COL_NAME - 1].value if len(row) >= COL_NAME else None
        email = row[COL_EMAIL - 1].value if len(row) >= COL_EMAIL else None
        category = row[COL_CATEGORY - 1].value if len(row) >= COL_CATEGORY else None
        emailed = row[COL_EMAILED - 1].value if len(row) >= COL_EMAILED else None

        leads.append({
            "row": row_idx,
            "name": name,
            "email": email,
            "category": category or "UNKNOWN",
            "emailed": emailed,
        })

    wb.close()
    return leads


def get_sendable_leads(leads, category_filter=None):
    """
    Filter leads to those that:
    1. Have an email address
    2. Have not been emailed yet
    3. Have a valid email format
    4. Match category filter (if set)
    """
    sendable = []
    for lead in leads:
        # Skip already emailed
        if lead["emailed"]:
            continue
        # Skip no email
        if not lead["email"]:
            continue
        # Category filter
        if category_filter and category_filter != "ALL":
            if lead["category"] != category_filter:
                continue
        # Validate email
        validated = validate_email(lead["email"])
        if not validated:
            continue

        lead["validated_email"] = validated
        sendable.append(lead)

    return sendable


def mark_as_emailed(row_number):
    """Update Excel to mark a row as emailed with current timestamp (column J)."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Ensure header exists
    header_cell = ws.cell(row=1, column=COL_EMAILED)
    if not header_cell.value:
        header_cell.value = "Emailed"

    # Write timestamp
    ws.cell(row=row_number, column=COL_EMAILED,
            value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    wb.save(EXCEL_FILE)
    wb.close()

# ─── Email Sending ────────────────────────────────────────────────────────────
def send_email(smtp_conn, from_addr, to_addr, subject, body, dry_run=False):
    """
    Send an email via an existing SMTP connection.
    Returns True on success, False on failure.
    """
    if dry_run:
        return True

    try:
        msg = MIMEMultipart()
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))

        smtp_conn.sendmail(from_addr, to_addr, msg.as_string())
        return True
    except Exception as e:
        logging.getLogger("email_sender").error(f"Failed to send to {to_addr}: {e}")
        return False


def connect_smtp(email_address, email_password):
    """
    Establish Gmail SMTP connection with TLS.
    Returns the SMTP connection object.
    """
    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(email_address, email_password)
    return server

# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Send emails to clinic leads")
    parser.add_argument("--dry-run", "-d", action="store_true",
                        help="Simulate sending without actually sending emails")
    parser.add_argument("--limit", "-l", type=int, default=0,
                        help="Maximum number of emails to send (0 = no limit)")
    parser.add_argument("--delay", type=int, default=DEFAULT_DELAY,
                        help=f"Seconds between emails (default: {DEFAULT_DELAY})")
    parser.add_argument("--category", "-c",
                        choices=["DENTAL", "HAIR_TRANSPLANT", "UNKNOWN", "ALL"],
                        default="ALL",
                        help="Send only to a specific category (default: ALL)")
    args = parser.parse_args()

    logger = setup_logging()
    mode = "DRY RUN" if args.dry_run else "LIVE"

    # ── Load credentials from environment ──
    email_address = os.environ.get("EMAIL_ADDRESS", "")
    email_password = os.environ.get("EMAIL_PASSWORD", "")

    if not args.dry_run and (not email_address or not email_password):
        print("\n  ERROR: Set EMAIL_ADDRESS and EMAIL_PASSWORD environment variables.")
        print("  Example:")
        print('    export EMAIL_ADDRESS="yourname@gmail.com"')
        print('    export EMAIL_PASSWORD="abcd efgh ijkl mnop"')
        print("\n  Or use --dry-run to simulate without credentials.\n")
        sys.exit(1)

    print()
    print("═══════════════════════════════════════════════════")
    print("  EMAIL SENDER")
    print(f"  Mode: {mode}")
    print(f"  From: {email_address or '(dry-run, no credentials needed)'}")
    print(f"  File: {EXCEL_FILE}")
    if args.limit:
        print(f"  Limit: {args.limit} emails")
    if args.category != "ALL":
        print(f"  Category: {args.category}")
    print(f"  Delay: {args.delay}s between emails")
    print("═══════════════════════════════════════════════════")
    print()

    # Load Airbnb info
    airbnb_info = load_airbnb_info()
    if airbnb_info:
        print(f"  Airbnb info: loaded from {AIRBNB_FILE.name}")
    else:
        print(f"  WARNING: {AIRBNB_FILE} not found!")
        print(f"  Create this file with your Airbnb property details.")
        print(f"  Emails will be sent without Airbnb details.")
        airbnb_info = "(Detajet e apartamentit do te shtohen se shpejti)"
    print()

    # Load and filter leads
    all_leads = load_leads()
    with_email = [l for l in all_leads if l["email"]]
    already_emailed = [l for l in all_leads if l["emailed"]]
    sendable = get_sendable_leads(all_leads, category_filter=args.category)

    print(f"  Total leads:        {len(all_leads)}")
    print(f"  With email:         {len(with_email)}")
    print(f"  Already emailed:    {len(already_emailed)}")
    print(f"  Sendable:           {len(sendable)}")

    if args.limit:
        sendable = sendable[:args.limit]
        print(f"  To send this run:   {len(sendable)} (limit)")
    else:
        print(f"  To send this run:   {len(sendable)}")
    print()

    if not sendable:
        print("  No leads to send emails to.")
        return

    # ── Connect to Gmail SMTP (once, reuse for all emails) ──
    smtp_conn = None
    if not args.dry_run:
        try:
            print("  Connecting to Gmail SMTP...")
            smtp_conn = connect_smtp(email_address, email_password)
            print("  Connected successfully.")
            print()
        except Exception as e:
            print(f"  ERROR: Failed to connect to Gmail SMTP: {e}")
            print("  Check your EMAIL_ADDRESS and EMAIL_PASSWORD.")
            print("  Make sure you're using a Gmail App Password, not your regular password.")
            sys.exit(1)

    # Sending loop
    sent = 0
    failed = 0

    try:
        for idx, lead in enumerate(sendable, start=1):
            name = clean_name(lead["name"])
            to_email = lead["validated_email"]
            category = lead["category"]
            subject = SUBJECTS.get(category, SUBJECTS["UNKNOWN"])
            template = MESSAGES.get(category, MESSAGES["UNKNOWN"])
            body = template.format(name=name, airbnb_info=airbnb_info)

            print(f"  [{idx}/{len(sendable)}] {name} | {to_email} | {category}")

            if args.dry_run:
                print(f"        [DRY RUN] Would send email ({len(body)} chars)")
                logger.info(f"DRY RUN | {to_email} | {name} | {category}")
                sent += 1
            else:
                logger.info(f"Sending to {to_email} | {name} | {category}")
                success = send_email(smtp_conn, email_address, to_email, subject, body)

                if success:
                    mark_as_emailed(lead["row"])
                    sent += 1
                    logger.info(f"SUCCESS | {to_email} | {name}")
                    print(f"        Sent successfully")
                else:
                    failed += 1
                    print(f"        FAILED — see log for details")

            sys.stdout.flush()

            # Delay between emails (skip after last one)
            if idx < len(sendable):
                if not args.dry_run:
                    print(f"        Waiting {args.delay}s...")
                    time.sleep(args.delay)
                else:
                    time.sleep(0.1)

            # Progress report every 10 emails
            if idx % 10 == 0:
                print(f"  ── Progress: {sent} sent, {failed} failed, "
                      f"{len(sendable) - idx} remaining")
                print()

    finally:
        # Always close SMTP connection
        if smtp_conn:
            try:
                smtp_conn.quit()
            except Exception:
                pass

    # Final summary
    print()
    print("═══════════════════════════════════════════════════")
    print("  COMPLETE")
    print(f"  Sent: {sent} | Failed: {failed}")
    if args.dry_run:
        print("  Mode: DRY RUN (no emails were actually sent)")
    print("═══════════════════════════════════════════════════")


if __name__ == "__main__":
    main()
