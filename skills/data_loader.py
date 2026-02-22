"""
Skill: Data Loader
Loads and parses all available data sources:
  - leads.xlsx (clinic lead database)
  - email_sender.log (email outreach history)
  - property_profile.md (property metadata)
  - airbnb_info.txt (property details)
  - marketing_report.md (existing analysis)

Returns a unified data dictionary for other skills to consume.
"""

import re
from datetime import datetime
from pathlib import Path

SKILL_NAME = "data_loader"
DESCRIPTION = "Loads all business data from Excel, logs, and config files"


def _find_project_root():
    """Find project root by looking for leads.xlsx relative to this file."""
    return Path(__file__).resolve().parent.parent


def load_leads(project_root=None):
    """
    Load all leads from leads.xlsx.
    Returns list of dicts with all columns.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    root = project_root or _find_project_root()
    excel_path = root / "leads.xlsx"

    if not excel_path.exists():
        return {"error": f"leads.xlsx not found at {excel_path}"}

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    leads = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        leads.append({
            "row": row_idx,
            "name": row[0] if len(row) > 0 else None,
            "location": row[1] if len(row) > 1 else None,
            "phone": row[2] if len(row) > 2 else None,
            "email": row[3] if len(row) > 3 else None,
            "category": row[4] if len(row) > 4 else None,
            "website": row[5] if len(row) > 5 else None,
            "date_added": str(row[6]) if len(row) > 6 and row[6] else None,
            "score": row[7] if len(row) > 7 else None,
            "messaged": str(row[8]) if len(row) > 8 and row[8] else None,
            "emailed": str(row[9]) if len(row) > 9 and row[9] else None,
        })

    wb.close()
    return leads


def load_email_log(project_root=None):
    """
    Parse email_sender.log into structured entries.
    Returns list of dicts with timestamp, level, action, email, name, category.
    """
    root = project_root or _find_project_root()
    log_path = root / "email_sender.log"

    if not log_path.exists():
        return {"error": f"email_sender.log not found at {log_path}"}

    entries = []
    pattern = re.compile(
        r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+) \| (\w+) \| (.+)$'
    )

    with open(log_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            m = pattern.match(line)
            if not m:
                continue

            timestamp_str, level, message = m.groups()
            entry = {
                "timestamp": timestamp_str,
                "level": level,
                "message": message,
            }

            # Parse specific message types
            if message.startswith("DRY RUN"):
                parts = message.split(" | ")
                if len(parts) >= 3:
                    entry["action"] = "dry_run"
                    entry["email"] = parts[1].strip()
                    entry["name"] = parts[2].strip()
                    if len(parts) >= 4:
                        entry["category"] = parts[3].strip()

            elif message.startswith("Sending to"):
                parts = message.replace("Sending to ", "").split(" | ")
                if len(parts) >= 2:
                    entry["action"] = "sending"
                    entry["email"] = parts[0].strip()
                    entry["name"] = parts[1].strip()
                    if len(parts) >= 3:
                        entry["category"] = parts[2].strip()

            elif message.startswith("SUCCESS"):
                parts = message.replace("SUCCESS | ", "").split(" | ")
                if len(parts) >= 2:
                    entry["action"] = "success"
                    entry["email"] = parts[0].strip()
                    entry["name"] = parts[1].strip()

            elif message.startswith("Failed"):
                entry["action"] = "failed"

            entries.append(entry)

    return entries


def load_property_profile(project_root=None):
    """Load property_profile.md as raw text."""
    root = project_root or _find_project_root()
    path = root / "property_profile.md"
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8")


def load_airbnb_info(project_root=None):
    """Load airbnb_info.txt as raw text."""
    root = project_root or _find_project_root()
    path = root / "airbnb_info.txt"
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8")


def load_marketing_report(project_root=None):
    """Load marketing_report.md as raw text."""
    root = project_root or _find_project_root()
    path = root / "marketing_report.md"
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8")


def load_property_fixes(project_root=None):
    """Load property_fixes.md as raw text."""
    root = project_root or _find_project_root()
    path = root / "property_fixes.md"
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8")


def run():
    """
    Main skill entry point.
    Loads all data sources and returns a summary.
    """
    root = _find_project_root()

    leads = load_leads(root)
    email_log = load_email_log(root)
    property_profile = load_property_profile(root)
    airbnb_info = load_airbnb_info(root)

    # Build summary
    if isinstance(leads, list):
        lead_count = len(leads)
        with_phone = sum(1 for l in leads if l["phone"])
        with_email = sum(1 for l in leads if l["email"])
        emailed = sum(1 for l in leads if l["emailed"])
        messaged = sum(1 for l in leads if l["messaged"])
        categories = {}
        for l in leads:
            cat = l["category"] or "UNKNOWN"
            categories[cat] = categories.get(cat, 0) + 1
    else:
        lead_count = 0
        with_phone = with_email = emailed = messaged = 0
        categories = {}

    if isinstance(email_log, list):
        log_count = len(email_log)
        successes = sum(1 for e in email_log if e.get("action") == "success")
        failures = sum(1 for e in email_log if e.get("action") == "failed")
    else:
        log_count = successes = failures = 0

    result = {
        "skill": SKILL_NAME,
        "status": "success",
        "data_sources": {
            "leads_xlsx": "loaded" if isinstance(leads, list) else "error",
            "email_log": "loaded" if isinstance(email_log, list) else "error",
            "property_profile": "loaded" if property_profile else "missing",
            "airbnb_info": "loaded" if airbnb_info else "missing",
        },
        "summary": {
            "total_leads": lead_count,
            "leads_with_phone": with_phone,
            "leads_with_email": with_email,
            "leads_emailed": emailed,
            "leads_messaged": messaged,
            "categories": categories,
            "email_log_entries": log_count,
            "email_successes": successes,
            "email_failures": failures,
        },
    }

    print(f"\n{'='*60}")
    print(f"  SKILL: {SKILL_NAME}")
    print(f"{'='*60}")
    print(f"  Total leads:       {lead_count}")
    print(f"  With phone:        {with_phone}")
    print(f"  With email:        {with_email}")
    print(f"  Already emailed:   {emailed}")
    print(f"  Already messaged:  {messaged}")
    print(f"  Categories:        {categories}")
    print(f"  Email log entries: {log_count}")
    print(f"  Email successes:   {successes}")
    print(f"  Email failures:    {failures}")
    print(f"{'='*60}\n")

    return result


if __name__ == "__main__":
    run()
